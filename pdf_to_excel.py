import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

PDF_PATH = "_INFO_RISCHI_SERVICE_IT_ENG_Scm.pdf"
OUTPUT_EXCEL = "modulo_sicurezza_layout.xlsx"


def format_sheet(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


def pdf_to_layout_excel(pdf_path, output_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modulo Sicurezza"

    current_row = 1

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = page.extract_text().split("\n")

            for line in lines:
                # Titoli di sezione
                if line.isupper() and len(line) > 10:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=4)
                    cell = ws.cell(row=current_row, column=1, value=line)
                    cell.font = Font(bold=True)
                    current_row += 1
                    continue

                # Domande numerate (es. 1.1, 2.3 ecc.)
                if line[:3].replace(".", "").isdigit():
                    parts = line.split(" ", 1)
                    ws.cell(row=current_row, column=1, value=parts[0])
                    ws.cell(row=current_row, column=2, value=parts[1] if len(parts) > 1 else "")
                    ws.cell(row=current_row, column=3, value="Y / N")
                    current_row += 1
                    continue

                # Testo normale
                ws.merge_cells(start_row=current_row, start_column=2,
                               end_row=current_row, end_column=4)
                ws.cell(row=current_row, column=2, value=line)
                current_row += 1

            current_row += 2  # spazio tra pagine

    format_sheet(ws)
    wb.save(output_excel)
    print(f"✓ Excel strutturato come il PDF: {output_excel}")


if __name__ == "__main__":
    pdf_to_layout_excel(PDF_PATH, OUTPUT_EXCEL)
